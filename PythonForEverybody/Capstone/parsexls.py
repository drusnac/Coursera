import sqlite3
import time
import ssl
import urllib.request, urllib.parse, urllib.error
from urllib.parse import urljoin
from urllib.parse import urlparse
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
#workbook = load_workbook('globalterrorism_small.xlsx', read_only=True)
workbook = load_workbook('globalterrorism_diluted.xlsx', read_only=True)

workSheet = workbook['Data']
#print(wb2)
workbookName = workbook.get_sheet_names()
print(workbookName)

#SQLITE Create Empty Database
conn = sqlite3.connect('terror_content.sqlite')
cur = conn.cursor()
cur.execute('''CREATE TABLE IF NOT EXISTS Terrorism2
    (id INTEGER UNIQUE, eventID TEXT, eventYear TEXT,
     countryNum TEXT, countryID TEXT, countryText TEXT)''')

countryIDs = {
"Afghanistan": "AFG",
"Angola": "AGO",
"Albania": "ALB",
"United Arab Emirates": "ARE",
"Argentina": "ARG",
"Armenia": "ARM",
"Antarctica": "ATA",
"French Southern and Antarctic Lands": "ATF",
"Australia": "AUS",
"Austria": "AUT",
"Azerbaijan": "AZE",
"Burundi": "BDI",
"Belgium": "BEL",
"Benin": "BEN",
"Burkina Faso": "BFA",
"Bangladesh": "BGD",
"Bulgaria": "BGR",
"The Bahamas": "BHS",
"Bosnia and Herzegovina": "BIH",
"Belarus": "BLR",
"Belize": "BLZ",
"Bolivia": "BOL",
"Brazil": "BRA",
"Brunei": "BRN",
"Bhutan": "BTN",
"Botswana": "BWA",
"Central African Republic": "CAF",
"Canada": "CAN",
"Switzerland": "CHE",
"Chile": "CHL",
"China": "CHN",
"Hong Kong": "CHN",
"Ivory Coast": "CIV",
"Cameroon": "CMR",
"Democratic Republic of the Congo": "COD",
"People's Republic of the Congo": "COD",
"Zaire": "COD",
"Republic of the Congo": "COG",
"Colombia": "COL",
"Costa Rica": "CRI",
"Cuba": "CUB",
"Northern Cyprus": "-99",
"Cyprus": "CYP",
"Czech Republic": "CZE",
"Germany": "DEU",
"West Germany (FRG)": "DEU",
"East Germany (GDR)": "DEU",
"Djibouti": "DJI",
"Denmark": "DNK",
"Dominican Republic": "DOM",
"Algeria": "DZA",
"Ecuador": "ECU",
"Egypt": "EGY",
"Eritrea": "ERI",
"Spain": "ESP",
"Estonia": "EST",
"Ethiopia": "ETH",
"Finland": "FIN",
"Fiji": "FJI",
"Falkland Islands": "FLK",
"France": "FRA",
"Andorra": "FRA",
"French Guiana": "GUF",
"Gabon": "GAB",
"United Kingdom": "GBR",
"Georgia": "GEO",
"Ghana": "GHA",
"Guinea": "GIN",
"Gambia": "GMB",
"Guinea Bissau": "GNB",
"Equatorial Guinea": "GNQ",
"Greece": "GRC",
"Greenland": "GRL",
"Guatemala": "GTM",
"Guyana": "GUY",
"Honduras": "HND",
"Croatia": "HRV",
"Haiti": "HTI",
"Hungary": "HUN",
"Indonesia": "IDN",
"India": "IND",
"Ireland": "IRL",
"Iran": "IRN",
"Iraq": "IRQ",
"Iceland": "ISL",
"Israel": "ISR",
"West Bank and Gaza Strip": "ISR",
"Italy": "ITA",
"Vatican City": "ITA",
"Jamaica": "JAM",
"Jordan": "JOR",
"Japan": "JPN",
"Kazakhstan": "KAZ",
"Kenya": "KEN",
"Kyrgyzstan": "KGZ",
"Cambodia": "KHM",
"South Korea": "KOR",
"Kosovo": "-99",
"Kuwait": "KWT",
"Laos": "LAO",
"Lebanon": "LBN",
"Liberia": "LBR",
"Libya": "LBY",
"Sri Lanka": "LKA",
"Lesotho": "LSO",
"Lithuania": "LTU",
"Luxembourg": "LUX",
"Latvia": "LVA",
"Morocco": "MAR",
"Moldova": "MDA",
"Madagascar": "MDG",
"Mexico": "MEX",
"Macedonia": "MKD",
"Mali": "MLI",
"Myanmar": "MMR",
"Montenegro": "MNE",
"Mongolia": "MNG",
"Mozambique": "MOZ",
"Mauritania": "MRT",
"Malawi": "MWI",
"Malaysia": "MYS",
"Singapore": "MYS",
"Namibia": "NAM",
"New Caledonia": "NCL",
"Niger": "NER",
"Nigeria": "NGA",
"Nicaragua": "NIC",
"Netherlands": "NLD",
"Norway": "NOR",
"Nepal": "NPL",
"New Zealand": "NZL",
"Oman": "OMN",
"Pakistan": "PAK",
"Panama": "PAN",
"Peru": "PER",
"Philippines": "PHL",
"Papua New Guinea": "PNG",
"Poland": "POL",
"Puerto Rico": "PRI",
"North Korea": "PRK",
"Portugal": "PRT",
"Paraguay": "PRY",
"Qatar": "QAT",
"Romania": "ROU",
"Russia": "RUS",
"Soviet Union": "RUS",
"Rwanda": "RWA",
"Western Sahara": "ESH",
"Saudi Arabia": "SAU",
"Sudan": "SDN",
"South Sudan": "SSD",
"Senegal": "SEN",
"Solomon Islands": "SLB",
"Sierra Leone": "SLE",
"El Salvador": "SLV",
"Somaliland": "-99",
"Somalia": "SOM",
"Republic of Serbia": "SRB",
"Yugoslavia": "SRB",
"Czechoslovakia": "SRB",
"Suriname": "SUR",
"Slovakia": "SVK",
"Slovenia": "SVN",
"Sweden": "SWE",
"Swaziland": "SWZ",
"Syria": "SYR",
"Chad": "TCD",
"Togo": "TGO",
"Thailand": "THA",
"Tajikistan": "TJK",
"Turkmenistan": "TKM",
"East Timor": "TLS",
"Trinidad and Tobago": "TTO",
"Tunisia": "TUN",
"Turkey": "TUR",
"Taiwan": "TWN",
"United Republic of Tanzania": "TZA",
"Tanzania": "TZA",
"Uganda": "UGA",
"Ukraine": "UKR",
"Uruguay": "URY",
"United States of America": "USA",
"United States": "USA",
"Uzbekistan": "UZB",
"Venezuela": "VEN",
"Vietnam": "VNM",
"South Vietnam": "VNM",
"Vanuatu": "VUT",
"West Bank": "PSE",
"Yemen": "YEM",
"North Yemen": "YEM",
"South Yemen": "YEM",
"South Africa": "ZAF",
"Zambia": "ZMB",
"Zimbabwe": "ZWE",
"Rhodesia": "ZWE",
              }
#PROCESS ROWS to get incident year and country
count = 0
for row in workSheet.rows:
    eventID = row[0].value
    eventYear = row[1].value
    countryNum = row[7].value
    countryText = row[8].value
    print(eventID)
    print(eventYear)
    print(countryNum)
    print(countryText)
    try:
        cur.execute('''INSERT OR IGNORE INTO Terrorism2 (id, eventID, eventYear, countryNum, countryID, countryText)
            VALUES ( ?, ?, ?, ?, ?, ? )''', (count, eventID, eventYear, countryNum, countryIDs[countryText] , countryText))
        count = count + 1
    except KeyError:
        print("Skipping " + countryText)
    conn.commit()
cur.close()






